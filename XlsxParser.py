# -*- coding:utf-8 -*-

from openpyxl import load_workbook

__author__ = "lqs"

class XLSXParser:
    def parse(self, fp, sheet_name=None):
        sheets = self.get_sheets(fp, sheet_name)
        for sheet in sheets:
            yield from self.extract_data_from_sheet(sheet)

    def extract_data_from_sheet(self, sheet):
        """Extract data from the sheet

        :param sheet: sheet name
        :return:

        We suppose the first row of the sheet is the row with the column titles, and the title cell
        can't be empty.
        """
        rows = sheet.max_row
        columns = sheet.max_column

        # extract data from the sheet starting on the second row
        for r in range(2, rows+1):
            yield {self.get_cell_value(sheet, 1, c): self.get_cell_value(sheet, r, c)
                   for c in range(1, columns+1)}

    @staticmethod
    def get_sheets(xlsx_name, sheet_name):
        """Extract all sheets in the excel file

        :param xlsx: excel file
        :param sheet_name: sheet name
        :return:

        An excel file may contain more than one sheet, and the construct between the sheets may be
        different, so we take the data extracted form every sheet as an item of the result list.
        """
        xlsx_book = load_workbook(xlsx_name)
        if sheet_name is None:
            sheets = [xlsx_book.get_sheet_by_name(sheet) for sheet in xlsx_book.get_sheet_names()]
        else:
            sheets = [xlsx_book.get_sheet_by_name(sheet_name)]

        return sheets

    @staticmethod
    def get_cell_value(sheet, row, column):
        """Extract data in the cell specified by row and column

        :param sheet: sheet name
        :param row: row number
        :param column: column number
        :return:
        """
        value = sheet.cell(row=row, column=column).value
        if value:
            return value.strip()

xp = XLSXParser()

if __name__ == '__main__':
    fp = '工作薄1.xlsx'
    for data in xp.parse(fp):
        if not data:
            continue
        print(data)
